from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from db import ROOT_DIR, connect, initialize, upsert_entry


RUNE_KIND_LABELS = {
    "WeaponRune": "무기 룬",
    "ArmorRune": "방어구 룬",
    "EmblemRune": "엠블럼 룬",
    "AccessoryRune": "장신구 룬",
}

TAG_EXPANSIONS = {
    "무방비": ["브레이크"],
}

EXCEL_SOURCES = ("db.xlsx#Runes", "db.xlsx#AccRune")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u200b", "").strip()


def split_list(value: Any) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def expand_tags(tags: list[str]) -> list[str]:
    expanded = list(tags)
    for tag in tags:
        expanded.extend(TAG_EXPANSIONS.get(tag, []))
    return list(dict.fromkeys(expanded))


def load_workbook(path: Path):
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl이 필요합니다. 설치: python -m pip install openpyxl") from exc

    return openpyxl.load_workbook(path, data_only=True)


def iter_rows_by_header(sheet):
    headers = [clean_text(cell.value) for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and clean_text(value) for value in row):
            continue
        yield dict(zip(headers, row))


def require_headers(sheet, expected: set[str]) -> None:
    headers = {clean_text(cell.value) for cell in sheet[1]}
    missing = sorted(expected - headers)
    if missing:
        raise SystemExit(f"{sheet.title} 시트에 필요한 컬럼이 없습니다: {', '.join(missing)}")


def delete_existing_excel_import(conn) -> None:
    placeholders = ", ".join("?" for _ in EXCEL_SOURCES)
    conn.execute(f"DELETE FROM entries WHERE source IN ({placeholders})", EXCEL_SOURCES)


def delete_orphan_tags(conn) -> None:
    conn.execute(
        """
        DELETE FROM tags
        WHERE NOT EXISTS (
            SELECT 1
            FROM entry_tags
            WHERE entry_tags.tag_id = tags.id
        )
        """
    )


def upsert_rune_detail(
    conn,
    *,
    entry_id: int,
    source_sheet: str,
    rune_kind: str,
    class_name: str = "",
    tier: str = "",
    skill_slot: int | None = None,
    raw_tags: str = "",
) -> None:
    conn.execute(
        """
        INSERT INTO rune_details(
            entry_id, source_sheet, rune_kind, class_name, tier, skill_slot, raw_tags
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(entry_id) DO UPDATE SET
            source_sheet = excluded.source_sheet,
            rune_kind = excluded.rune_kind,
            class_name = excluded.class_name,
            tier = excluded.tier,
            skill_slot = excluded.skill_slot,
            raw_tags = excluded.raw_tags
        """,
        (entry_id, source_sheet, rune_kind, class_name, tier, skill_slot, raw_tags),
    )


def import_runes_sheet(conn, sheet) -> int:
    require_headers(sheet, {"name", "type", "tier", "description", "tag"})

    count = 0
    for row in iter_rows_by_header(sheet):
        name = clean_text(row.get("name"))
        rune_kind = clean_text(row.get("type"))
        tier = clean_text(row.get("tier"))
        description = clean_text(row.get("description"))
        raw_tags = clean_text(row.get("tag"))
        if not name or not rune_kind:
            continue

        kind_label = RUNE_KIND_LABELS.get(rune_kind, rune_kind)
        raw_tag_list = split_list(raw_tags)
        tags = expand_tags(["룬", kind_label, rune_kind, *split_list(tier), *raw_tag_list])
        display_tags = ", ".join(expand_tags(raw_tag_list))
        entry_id = upsert_entry(
            conn,
            entry_type=rune_kind,
            name=name,
            summary=f"{kind_label} / {tier}",
            description=description,
            source="db.xlsx#Runes",
            tags=tags,
            attributes={
                "시트": "Runes",
                "룬 종류": kind_label,
                "룬 종류 코드": rune_kind,
                "등급": tier,
                "태그": display_tags,
            },
            replace_tags=True,
        )
        upsert_rune_detail(
            conn,
            entry_id=entry_id,
            source_sheet="Runes",
            rune_kind=rune_kind,
            tier=tier,
            raw_tags=raw_tags,
        )
        count += 1
    return count


def import_acc_rune_sheet(conn, sheet) -> int:
    require_headers(sheet, {"class", "name", "type", "tier", "description", "skill"})

    count = 0
    for row in iter_rows_by_header(sheet):
        class_name = clean_text(row.get("class"))
        name = clean_text(row.get("name"))
        rune_kind = clean_text(row.get("type")) or "AccessoryRune"
        tier = clean_text(row.get("tier"))
        description = clean_text(row.get("description"))
        skill_text = clean_text(row.get("skill"))
        skill_slot = int(float(skill_text)) if skill_text else None
        if not class_name or not name:
            continue

        kind_label = RUNE_KIND_LABELS.get(rune_kind, rune_kind)
        skill_tag = f"스킬 {skill_slot}" if skill_slot is not None else ""
        tags = ["룬", kind_label, rune_kind, "AccRune", class_name, *split_list(tier)]
        if skill_tag:
            tags.append(skill_tag)

        entry_id = upsert_entry(
            conn,
            entry_type=rune_kind,
            name=name,
            summary=f"{class_name} {kind_label} / {tier}" + (f" / {skill_tag}" if skill_tag else ""),
            description=description,
            source="db.xlsx#AccRune",
            tags=tags,
            attributes={
                "시트": "AccRune",
                "룬 종류": kind_label,
                "룬 종류 코드": rune_kind,
                "직업": class_name,
                "등급": tier,
                "스킬 슬롯": str(skill_slot) if skill_slot is not None else "",
            },
            replace_tags=True,
        )
        upsert_rune_detail(
            conn,
            entry_id=entry_id,
            source_sheet="AccRune",
            rune_kind=rune_kind,
            class_name=class_name,
            tier=tier,
            skill_slot=skill_slot,
        )
        count += 1
    return count


def import_excel(path: Path) -> tuple[int, int]:
    initialize()
    workbook = load_workbook(path)
    with connect() as conn:
        delete_existing_excel_import(conn)
        runes_count = import_runes_sheet(conn, workbook["Runes"])
        acc_count = import_acc_rune_sheet(conn, workbook["AccRune"])
        delete_orphan_tags(conn)
    return runes_count, acc_count


def main() -> None:
    parser = argparse.ArgumentParser(description="Import rune data from db.xlsx into SQLite.")
    parser.add_argument("--file", default=str(ROOT_DIR / "db.xlsx"), help="Excel file path")
    args = parser.parse_args()

    path = Path(args.file)
    runes_count, acc_count = import_excel(path)
    print(f"Imported Runes: {runes_count}")
    print(f"Imported AccRune: {acc_count}")


if __name__ == "__main__":
    main()
